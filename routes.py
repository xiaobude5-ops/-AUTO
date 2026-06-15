"""Flask 路由 — 页面 + API"""
import json
import threading
import time
from datetime import datetime
from io import BytesIO

from flask import request, jsonify, render_template, send_file, Response, session, redirect, url_for

from auth import (
    hash_password, verify_password,
    login_required, api_login_required, admin_required, api_admin_required,
)
from database import (
    create_user, get_user_by_username, get_user_by_id, get_all_users,
    update_user_role, update_user_password, delete_user, get_employees,
    add_submission, save_video_snapshot, get_submissions_by_user,
    get_submissions_filtered, get_submission, update_payment,
    delete_submission, update_submission_meta,
    get_daily_changes_for_user, get_admin_monthly_summary,
    get_all_submission_aweme_ids,
)
from scraper import scrape_single_video_with_retry, NotLoggedInError
from utils import extract_aweme_id, get_month_boundaries
from config import SCRAPE_RETRY_MAX, PORT


def register_routes(app):
    # ═══════════════════════════════════════════
    # 页面路由
    # ═══════════════════════════════════════════

    @app.route("/login")
    def login():
        return render_template("login.html")

    @app.route("/register")
    def register():
        return render_template("register.html")

    @app.route("/")
    @login_required
    def dashboard():
        return render_template("dashboard.html")

    @app.route("/admin")
    @admin_required
    def admin_panel():
        return render_template("admin.html")

    @app.route("/admin/users")
    @admin_required
    def admin_users():
        return render_template("admin.html", page="users")

    @app.route("/admin/monthly")
    @admin_required
    def admin_monthly():
        return render_template("admin_monthly.html")

    # ═══════════════════════════════════════════
    # 认证 API
    # ═══════════════════════════════════════════

    @app.route("/api/login", methods=["POST"])
    def api_login():
        data = request.get_json(force=True) or {}
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        user = get_user_by_username(username)
        if not user or not verify_password(password, user["password_hash"]):
            return jsonify({"success": False, "error": "用户名或密码错误"}), 401

        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["nickname"] = user["nickname"]
        session["role"] = user["role"]
        return jsonify({"success": True, "data": {
            "id": user["id"], "username": user["username"],
            "nickname": user["nickname"], "role": user["role"],
        }})

    @app.route("/api/register", methods=["POST"])
    def api_register():
        data = request.get_json(force=True) or {}
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""
        nickname = (data.get("nickname") or "").strip()

        if not username or len(username) < 2:
            return jsonify({"success": False, "error": "用户名至少2个字符"}), 400
        if not password or len(password) < 4:
            return jsonify({"success": False, "error": "密码至少4位"}), 400
        if not nickname:
            return jsonify({"success": False, "error": "请输入昵称"}), 400

        user_id = create_user(username, hash_password(password), nickname, "employee")
        if not user_id:
            return jsonify({"success": False, "error": "用户名已存在"}), 400

        return jsonify({"success": True, "data": {"id": user_id, "nickname": nickname}})

    @app.route("/api/logout", methods=["POST"])
    def api_logout():
        session.clear()
        return jsonify({"success": True})

    @app.route("/api/me")
    @api_login_required
    def api_me():
        return jsonify({"success": True, "data": {
            "id": session["user_id"],
            "username": session["username"],
            "nickname": session["nickname"],
            "role": session["role"],
        }})

    @app.route("/api/change-password", methods=["POST"])
    @api_login_required
    def api_change_password():
        data = request.get_json(force=True) or {}
        old_pw = data.get("old_password") or ""
        new_pw = data.get("new_password") or ""
        if not old_pw or not new_pw:
            return jsonify({"success": False, "error": "请填写新旧密码"}), 400
        if len(new_pw) < 4:
            return jsonify({"success": False, "error": "新密码至少4位"}), 400
        user = get_user_by_id(session["user_id"])
        if not user or not verify_password(old_pw, user["password_hash"]):
            return jsonify({"success": False, "error": "原密码错误"}), 400
        update_user_password(session["user_id"], hash_password(new_pw))
        return jsonify({"success": True})

    # ═══════════════════════════════════════════
    # 用户管理 API（管理员）
    # ═══════════════════════════════════════════

    @app.route("/api/admin/users")
    @api_admin_required
    def api_admin_users():
        return jsonify({"success": True, "data": get_all_users()})

    @app.route("/api/admin/users/<int:user_id>/role", methods=["PATCH"])
    @api_admin_required
    def api_admin_update_role(user_id):
        data = request.get_json(force=True) or {}
        role = data.get("role")
        if role not in ("admin", "employee"):
            return jsonify({"success": False, "error": "无效角色"}), 400
        # 保护初始 admin
        user = get_user_by_id(user_id)
        if user and user["username"] == "admin" and user["role"] == "admin" and role != "admin":
            return jsonify({"success": False, "error": "初始管理员不可降权"}), 400
        update_user_role(user_id, role)
        return jsonify({"success": True})

    @app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
    @api_admin_required
    def api_admin_delete_user(user_id):
        user = get_user_by_id(user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在"}), 404
        if user["username"] == "admin":
            return jsonify({"success": False, "error": "初始管理员不可删除"}), 400
        if user["id"] == session["user_id"]:
            return jsonify({"success": False, "error": "不可删除当前登录账号"}), 400
        delete_user(user_id)
        return jsonify({"success": True})

    @app.route("/api/employees")
    @api_login_required
    def api_employees():
        return jsonify({"success": True, "data": get_employees()})

    @app.route("/api/daily-changes")
    @api_login_required
    def api_daily_changes():
        changes = get_daily_changes_for_user(session["user_id"])
        return jsonify({"success": True, "data": changes or []})

    # ═══════════════════════════════════════════
    # 提交 API
    # ═══════════════════════════════════════════

    @app.route("/api/submissions", methods=["POST"])
    @api_login_required
    def api_submit():
        data = request.get_json(force=True) or {}
        url = (data.get("url") or "").strip()
        category = (data.get("category") or "").strip()
        collaborator_ids = data.get("collaborator_ids") or []

        if not url:
            return jsonify({"success": False, "error": "请提供视频链接"}), 400
        if category not in ("日常", "商单", "爆款", "金劫宝"):
            return jsonify({"success": False, "error": "分类必须是「日常/商单/爆款/金劫宝」"}), 400

        aweme_id = extract_aweme_id(url)
        if not aweme_id:
            return jsonify({"success": False, "error": "无法解析抖音视频链接"}), 400

        # 抓取
        try:
            video_data = scrape_single_video_with_retry(aweme_id, SCRAPE_RETRY_MAX)
        except NotLoggedInError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        if not video_data:
            return jsonify({"success": False, "error": "抓取失败，请检查链接是否正确"}), 400

        # 保存
        sid = add_submission(session["user_id"], aweme_id, url, category, collaborator_ids)
        if sid is None:
            return jsonify({"success": False, "error": "该视频已提交过"}), 400

        save_video_snapshot(sid, video_data["likes"], video_data["comments"],
                           video_data["shares"], video_data["collects"])

        return jsonify({"success": True, "data": {
            "id": sid,
            "aweme_id": aweme_id,
            "desc": video_data["desc"],
            "create_date": video_data["create_date"],
            "likes": video_data["likes"],
            "comments": video_data["comments"],
            "shares": video_data["shares"],
            "collects": video_data["collects"],
        }})

    @app.route("/api/submissions")
    @api_login_required
    def api_submissions():
        submissions = get_submissions_by_user(session["user_id"])
        return jsonify({"success": True, "data": submissions})

    @app.route("/api/submissions/<int:sid>")
    @api_login_required
    def api_submission_detail(sid):
        s = get_submission(sid)
        if not s:
            return jsonify({"success": False, "error": "不存在"}), 404
        if session["role"] != "admin" and s["user_id"] != session["user_id"]:
            return jsonify({"success": False, "error": "无权限"}), 403
        return jsonify({"success": True, "data": s})

    @app.route("/api/submissions/<int:sid>", methods=["PATCH"])
    @api_login_required
    def api_submission_update(sid):
        s = get_submission(sid)
        if not s:
            return jsonify({"success": False, "error": "不存在"}), 404
        if s["user_id"] != session["user_id"]:
            return jsonify({"success": False, "error": "无权限"}), 403

        data = request.get_json(force=True) or {}
        category = (data.get("category") or "").strip() or None
        submitted_at = (data.get("submitted_at") or "").strip() or None

        if category and category not in ("日常", "商单", "爆款", "金劫宝"):
            return jsonify({"success": False, "error": "无效分类"}), 400
        if submitted_at:
            import re
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", submitted_at):
                return jsonify({"success": False, "error": "日期格式需为 YYYY-MM-DD"}), 400

        update_submission_meta(sid, category=category, submitted_at=submitted_at)
        return jsonify({"success": True})

    # ═══════════════════════════════════════════
    # 管理员筛选 + 汇总 API
    # ═══════════════════════════════════════════

    @app.route("/api/admin/filter")
    @api_admin_required
    def api_admin_filter():
        user_ids_str = request.args.get("user_id", "")
        user_ids = [int(x) for x in user_ids_str.split(",") if x] if user_ids_str else None
        category = request.args.get("category")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        engagement_min = request.args.get("engagement_min", 0, type=int)
        payment_min = request.args.get("payment_min", type=float)
        payment_max = request.args.get("payment_max", type=float)
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)

        results, total = get_submissions_filtered(
            user_ids=user_ids, category=category,
            date_from=date_from, date_to=date_to,
            engagement_min=engagement_min,
            payment_min=payment_min, payment_max=payment_max,
            page=page, per_page=per_page,
        )

        return jsonify({"success": True, "data": results, "total": total})

    @app.route("/api/admin/payment/<int:sid>", methods=["PATCH"])
    @api_admin_required
    def api_admin_payment(sid):
        data = request.get_json(force=True) or {}
        amount = data.get("amount")
        update_payment(sid, amount)
        return jsonify({"success": True})

    @app.route("/api/admin/submissions/<int:sid>", methods=["DELETE"])
    @api_admin_required
    def api_admin_delete_submission(sid):
        if not get_submission(sid):
            return jsonify({"success": False, "error": "记录不存在"}), 404
        delete_submission(sid)
        return jsonify({"success": True})

    @app.route("/api/admin/daily-summary")
    @api_admin_required
    def api_admin_daily_summary():
        """管理员全局每日变化（按选定用户汇总）"""
        user_ids_str = request.args.get("user_id", "")
        if user_ids_str:
            user_ids = [int(x) for x in user_ids_str.split(",") if x]
        else:
            # 默认所有员工
            employees = get_employees()
            user_ids = [e["id"] for e in employees]

        all_changes = []
        for uid in user_ids:
            changes = get_daily_changes_for_user(uid)
            if changes:
                user = get_user_by_id(uid)
                for c in changes:
                    c["nickname"] = user["nickname"] if user else ""
                all_changes.extend(changes)

        all_changes.sort(key=lambda x: x["total_change"], reverse=True)
        return jsonify({"success": True, "data": all_changes})

    @app.route("/api/admin/monthly-summary")
    @api_admin_required
    def api_admin_monthly_summary():
        year = request.args.get("year", type=int)
        month = request.args.get("month", type=int)
        if not year or not month:
            now = datetime.now()
            year, month = now.year, now.month
        data = get_admin_monthly_summary(year, month)
        return jsonify({"success": True, "data": data})

    # ═══════════════════════════════════════════
    # Excel 导出
    # ═══════════════════════════════════════════

    @app.route("/api/admin/export")
    @api_admin_required
    def api_admin_export():
        year = request.args.get("year", type=int)
        month = request.args.get("month", type=int)
        user_ids_str = request.args.get("user_id", "")
        user_ids = [int(x) for x in user_ids_str.split(",") if x] if user_ids_str else None
        category = request.args.get("category")
        engagement_min = request.args.get("engagement_min", 0, type=int)
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")

        if not year or not month:
            now = datetime.now()
            year, month = now.year, now.month

        results, _ = get_submissions_filtered(
            user_ids=user_ids, category=category,
            date_from=date_from, date_to=date_to,
            engagement_min=engagement_min,
            page=1, per_page=10000,
        )

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d}"
        ws.sheet_view.showGridLines = False

        # ── 主题色（与前端 Dark Luxury 呼应，但 Excel 用浅色提高可读性）──
        GOLD = "C9A96E"
        GOLD_DARK = "8B6F2E"
        GOLD_BG_LIGHT = "FAF6EC"
        BG_HEADER = "1A1F35"     # 深海蓝紫
        TEXT_LIGHT = "FFFFFF"
        BG_ALT = "F7F4ED"        # 米色斑马
        BG_TOTAL = "EFE6CC"      # 合计行金米色
        BORDER_GOLD = "D4BE8E"

        thin = Side(border_style="thin", color="E2D5B5")
        thick_gold = Side(border_style="medium", color=GOLD)
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── 标题行 ──
        ws.merge_cells("A1:L1")
        c = ws["A1"]
        c.value = f"几何星球  ·  数据汇总报表"
        c.font = Font(name="微软雅黑", bold=True, size=20, color=GOLD_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=GOLD_BG_LIGHT)
        ws.row_dimensions[1].height = 38

        # ── 副标题行 ──
        ws.merge_cells("A2:L2")
        c = ws["A2"]
        total_eng = sum((p.get("likes") or 0) + (p.get("comments") or 0)
                        + (p.get("shares") or 0) + (p.get("collects") or 0) for p in results)
        total_pay = sum(p.get("payment_amount") or 0 for p in results)
        c.value = (f"{year}年{month}月    "
                   f"导出时间 {datetime.now().strftime('%Y-%m-%d %H:%M')}    "
                   f"共 {len(results)} 条记录    "
                   f"互动总量 {total_eng:,}    "
                   f"商单总额 ¥{total_pay:,.0f}")
        c.font = Font(name="微软雅黑", size=10, color="6B5C3A", italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=GOLD_BG_LIGHT)
        ws.row_dimensions[2].height = 22

        # ── 空白分隔（金色描边） ──
        ws.row_dimensions[3].height = 6
        for col in range(1, 13):
            ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=GOLD)

        # ── 表头 ──
        headers = ["#", "员工", "提交日期", "分类", "协作人员", "商单金额",
                   "👍 点赞", "💬 评论", "🔄 分享", "⭐ 收藏", "🔥 互动", "视频链接"]
        hdr_fill = PatternFill("solid", fgColor=BG_HEADER)
        hdr_font = Font(name="微软雅黑", bold=True, color=TEXT_LIGHT, size=11)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thick_gold, bottom=thick_gold, left=thin, right=thin)
        ws.row_dimensions[4].height = 32

        # ── 数据行 ──
        data_font = Font(name="微软雅黑", size=10, color="2A2520")
        num_font = Font(name="Consolas", size=10, color="2A2520")
        gold_font = Font(name="Consolas", size=10, bold=True, color=GOLD_DARK)
        link_font = Font(name="微软雅黑", size=9, color="6B7C8C", underline="single")

        for idx, p in enumerate(results, 1):
            row = 4 + idx
            eng = (p.get("likes") or 0) + (p.get("comments") or 0) + \
                  (p.get("shares") or 0) + (p.get("collects") or 0)
            collab_str = ", ".join(c["nickname"] for c in (p.get("collaborators") or []))
            payment = p.get("payment_amount")
            video_url = (
                f"https://www.douyin.com/video/{p['aweme_id']}" if p.get("aweme_id")
                else (p.get("video_url") or "")
            )

            row_bg = BG_ALT if idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill("solid", fgColor=row_bg)

            vals = [
                idx,
                p.get("submitter_name", ""),
                (p.get("submitted_at") or "")[:10],
                p.get("category", ""),
                collab_str,
                payment if payment else "",
                p.get("likes") or 0,
                p.get("comments") or 0,
                p.get("shares") or 0,
                p.get("collects") or 0,
                eng,
                video_url,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.fill = row_fill
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal=("right" if col >= 6 and col <= 11 else
                                "center" if col in (1, 3, 4) else "left"),
                    vertical="center",
                )
                if col == 1:
                    cell.font = Font(name="Consolas", size=10, color="A89878")
                elif col == 4:  # 分类颜色
                    cat_colors = {
                        "商单": GOLD_DARK, "爆款": "C0392B", "金劫宝": "8E44AD",
                    }
                    cell.font = Font(name="微软雅黑", size=9, bold=True,
                                     color=cat_colors.get(v, "4A6B8C"))
                elif col == 6:  # 金额
                    cell.font = gold_font
                    cell.number_format = '"¥"#,##0;[Red]-"¥"#,##0;"—"'
                elif col == 11:  # 互动总量
                    cell.font = Font(name="Consolas", size=10, bold=True, color=GOLD_DARK)
                    cell.number_format = "#,##0"
                elif col >= 7 and col <= 10:
                    cell.font = num_font
                    cell.number_format = "#,##0"
                elif col == 12:
                    cell.font = link_font
                    if video_url:
                        cell.hyperlink = video_url
                else:
                    cell.font = data_font
            ws.row_dimensions[row].height = 22

        # ── 合计行 ──
        tr = 4 + len(results) + 1
        ws.row_dimensions[tr].height = 28
        total_fill = PatternFill("solid", fgColor=BG_TOTAL)
        ws.merge_cells(start_row=tr, end_row=tr, start_column=1, end_column=5)
        c = ws.cell(row=tr, column=1, value="合  计")
        c.font = Font(name="微软雅黑", bold=True, size=11, color=GOLD_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = total_fill
        c.border = Border(top=thick_gold, bottom=thick_gold, left=thin, right=thin)

        sums = {
            "payment_amount": sum(p.get("payment_amount") or 0 for p in results),
            "likes":     sum(p.get("likes") or 0 for p in results),
            "comments":  sum(p.get("comments") or 0 for p in results),
            "shares":    sum(p.get("shares") or 0 for p in results),
            "collects":  sum(p.get("collects") or 0 for p in results),
        }
        sums["eng"] = sums["likes"] + sums["comments"] + sums["shares"] + sums["collects"]

        total_cells = [
            (6, sums["payment_amount"], '"¥"#,##0'),
            (7, sums["likes"], "#,##0"),
            (8, sums["comments"], "#,##0"),
            (9, sums["shares"], "#,##0"),
            (10, sums["collects"], "#,##0"),
            (11, sums["eng"], "#,##0"),
        ]
        for col, v, fmt in total_cells:
            cell = ws.cell(row=tr, column=col, value=v)
            cell.font = Font(name="Consolas", bold=True, size=11, color=GOLD_DARK)
            cell.number_format = fmt
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(top=thick_gold, bottom=thick_gold, left=thin, right=thin)
        # 链接列在合计行留空但保持样式
        cell = ws.cell(row=tr, column=12, value="")
        cell.fill = total_fill
        cell.border = Border(top=thick_gold, bottom=thick_gold, left=thin, right=thin)

        # ── 列宽 ──
        widths = [6, 12, 13, 8, 22, 14, 11, 11, 11, 11, 13, 32]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── 冻结表头 + 自动筛选 ──
        ws.freeze_panes = "A5"
        if results:
            ws.auto_filter.ref = f"A4:L{4 + len(results)}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"几何星球_数据汇总_{year}{month:02d}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    # ═══════════════════════════════════════════
    # 刷新 API
    # ═══════════════════════════════════════════

    @app.route("/api/refresh", methods=["POST"])
    @api_admin_required
    def api_refresh():
        """管理员手动触发全局刷新"""
        t = threading.Thread(target=_refresh_all_background, args=(app,), daemon=True)
        t.start()
        return jsonify({"success": True, "data": {"status": "started"}})

    app._refresh_running = False
    app._refresh_errors: list[str] = []

    def _refresh_all_background(app):
        if app._refresh_running:
            return
        app._refresh_running = True
        app._refresh_errors = []
        try:
            items = get_all_submission_aweme_ids()
            for item in items:
                for attempt in range(SCRAPE_RETRY_MAX + 1):
                    try:
                        data = scrape_single_video_with_retry(item["aweme_id"])
                        if data:
                            save_video_snapshot(item["id"], data["likes"], data["comments"],
                                               data["shares"], data["collects"])
                        break
                    except NotLoggedInError as e:
                        app._refresh_errors.append(str(e))
                        return  # 全部中断 — 登录态丢失，重试无意义
                    except Exception as e:
                        if attempt >= SCRAPE_RETRY_MAX:
                            app._refresh_errors.append(f"{item['aweme_id']}: {e}")
                time.sleep(1)
        finally:
            app._refresh_running = False
